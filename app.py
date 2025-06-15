import tkinter as tk
from tkinter import ttk
import pandas as pd
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from copy import copy

load_dotenv()
# Checking
# Load config from .env
excel_file = os.getenv("EXCEL_FILE")
columns_to_display = os.getenv("COLUMNS_TO_DISPLAY").split(",")
window_size = os.getenv("WINDOW_SIZE")
output_excel_file = os.getenv("OUTPUT_EXCEL_FILE", "quote_output.xlsx")
template_file = os.getenv("TEMPLATE_FILE", "template.xlsx")

# Load data from Excel
df = pd.read_excel(excel_file, skiprows=5)
df["Part Name"] = df["Part Name"].astype(str)

saved_products = []

root = tk.Tk()
root.geometry(window_size)
root.title("Price List Quoter")
root.state('zoomed')

main_frame = ttk.Frame(root, padding=20)
main_frame.pack(fill="both", expand=True)

left_frame = ttk.Frame(main_frame)
left_frame.pack(side="left", fill="both", expand=True, padx=10)

right_frame = ttk.Frame(main_frame)
right_frame.pack(side="right", fill="both", expand=True, padx=10)

# def create_labeled_entry(parent, label_text, var, readonly=False):
#     ttk.Label(parent, text=label_text, font=("Arial", 12)).pack(anchor="w", pady=(10, 0))
#     entry = ttk.Entry(parent, textvariable=var)
#     if readonly:
#         entry.config(state="readonly")
#     entry.pack(fill="x")
#     return entry

# Utility function to clone cell style
def copy_cell_style(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)

def create_labeled_entry(parent, label_text, var, readonly=False):
    ttk.Label(parent, text=label_text, font=("Arial", 12)).pack(anchor="w", pady=(10, 0))
    entry = ttk.Entry(parent, textvariable=var)
    if readonly:
        entry.config(state="readonly")
    entry.pack(fill="x")
    return entry

entry_var = tk.StringVar()
description_var = tk.StringVar()
list_price_var = tk.DoubleVar(value=0.0)
qty_var = tk.IntVar(value=1)
discount_var = tk.IntVar(value=0)
unit_price_var = tk.DoubleVar(value=0.0)
extended_price_var = tk.DoubleVar(value=0.0)
margin_var = tk.IntVar(value=0)
margin_unit_price_var = tk.DoubleVar(value=0.0)
margin_extended_price_var = tk.DoubleVar(value=0.0)

# Part Name input
ttk.Label(left_frame, text="Part Name:", font=("Arial", 12)).pack(anchor="w", pady=(0, 2))
entry = ttk.Entry(left_frame, textvariable=entry_var, width=40)
entry.pack(anchor="w", fill="x")

listbox_frame = ttk.Frame(left_frame)
listbox_frame.pack(fill="x")
scrollbar = ttk.Scrollbar(listbox_frame, orient="vertical")
listbox = tk.Listbox(listbox_frame, height=5, yscrollcommand=scrollbar.set, exportselection=False)
scrollbar.config(command=listbox.yview)
scrollbar.pack(side="right", fill="y")
listbox.pack(side="left", fill="x", expand=True)

create_labeled_entry(left_frame, "Qty", qty_var)
create_labeled_entry(left_frame, "Discount (%)", discount_var)
create_labeled_entry(left_frame, "Margin (%)", margin_var)

create_labeled_entry(right_frame, "Description", description_var, readonly=True)
create_labeled_entry(right_frame, "List Price (RO)", list_price_var, readonly=True)
create_labeled_entry(right_frame, "Unit Price (RO)", unit_price_var, readonly=True)
create_labeled_entry(right_frame, "Extended Price (RO)", extended_price_var, readonly=True)
create_labeled_entry(right_frame, "Margin Unit Price (RO)", margin_unit_price_var, readonly=True)
create_labeled_entry(right_frame, "Margin Extended Price (RO)", margin_extended_price_var, readonly=True)

def update_listbox(*args):
    listbox.delete(0, tk.END)
    typed = entry_var.get().lower()
    if typed:
        matches = df[df["Part Name"].str.lower().str.contains(typed, na=False)]["Part Name"].unique()
        for match in matches:
            listbox.insert(tk.END, match)

entry_var.trace_add("write", update_listbox)

selected_index = -1

def select_from_listbox(index):
    if 0 <= index < listbox.size():
        listbox.selection_clear(0, tk.END)
        listbox.selection_set(index)
        listbox.activate(index)
        listbox.see(index)

def on_key_press(event):
    global selected_index
    if event.keysym == "Down":
        if listbox.size() > 0:
            selected_index = min(selected_index + 1, listbox.size() - 1)
            select_from_listbox(selected_index)
    elif event.keysym == "Up":
        if listbox.size() > 0:
            selected_index = max(selected_index - 1, 0)
            select_from_listbox(selected_index)
    elif event.keysym == "Return":
        if 0 <= selected_index < listbox.size():
            selected = listbox.get(selected_index)
            entry_var.set(selected)
            match = df[df["Part Name"] == selected].iloc[0]
            description_var.set(match["Description"])
            list_price_var.set(match["List Price"])
            recalculate()
            listbox.delete(0, tk.END)
            selected_index = -1

entry.bind("<Down>", on_key_press)
entry.bind("<Up>", on_key_press)
entry.bind("<Return>", on_key_press)

listbox.bind("<<ListboxSelect>>", lambda e: on_listbox_select())

def on_listbox_select():
    global selected_index
    if listbox.curselection():
        selected_index = listbox.curselection()[0]
        selected = listbox.get(selected_index)
        entry_var.set(selected)
        match = df[df["Part Name"] == selected].iloc[0]
        description_var.set(match["Description"])
        list_price_var.set(match["List Price"])
        recalculate()
        listbox.delete(0, tk.END)
        selected_index = -1

def recalculate(*args):
    try:
        qty = max(qty_var.get(), 0)
        try:
            list_price = float(list_price_var.get())
        except (ValueError, TypeError):
            list_price = 0.0

        discount = min(max(discount_var.get(), 0), 100)
        margin = min(max(margin_var.get(), 0), 100)

        unit_price = list_price * (1 - discount / 100)
        extended_price = unit_price * qty
        margin_unit_price = unit_price * (1 + margin / 100)
        margin_extended_price = margin_unit_price * qty

        unit_price_var.set(round(unit_price, 2))
        extended_price_var.set(round(extended_price, 2))
        margin_unit_price_var.set(round(margin_unit_price, 2))
        margin_extended_price_var.set(round(margin_extended_price, 2))

    except Exception as e:
        print(f"Calculation error: {e}")

qty_var.trace_add("write", recalculate)
discount_var.trace_add("write", recalculate)
margin_var.trace_add("write", recalculate)

def clear_all():
    entry_var.set("")
    description_var.set("")
    list_price_var.set(0.0)
    qty_var.set(1)
    discount_var.set(0)
    unit_price_var.set(0.0)
    extended_price_var.set(0.0)
    margin_var.set(0)
    margin_unit_price_var.set(0.0)
    margin_extended_price_var.set(0.0)
    listbox.delete(0, tk.END)
    global selected_index
    selected_index = -1

def save_product():
    if not entry_var.get():
        print("Part Name is required.")
        return
    data = {
        "Part Name": entry_var.get(),
        "Description": description_var.get(),
        "Qty": qty_var.get(),
        "List Price": list_price_var.get(),
        "Discount": discount_var.get(),
        "Unit Price": unit_price_var.get(),
        "Extended Price": extended_price_var.get(),
        "Margin": margin_var.get(),
        "Margin Unit Price": margin_unit_price_var.get(),
        "Margin Extended Price": margin_extended_price_var.get()
    }
    saved_products.append(data)
    print(f"Saved: {data['Part Name']}")
    clear_all()

def export_to_excel():
    if not saved_products:
        print("No products to export.")
        return
    try:
        df_export = pd.DataFrame(saved_products)
        df_export.to_excel(output_excel_file, index=False)
        print(f"Exported all saved products to '{output_excel_file}'")
        saved_products.clear()
    except Exception as e:
        print(f"Error exporting: {e}")

def export_to_final_excel():
    if not os.path.exists(output_excel_file):
        print("Excel file not found. Export to Excel first.")
        return

    try:
        df_export = pd.read_excel(output_excel_file)
        wb_template = load_workbook("Header.xlsx")
        ws_template = wb_template.active
    except Exception as e:
        print(f"Failed to read Excel files: {e}")
        return

    if df_export.empty:
        print("No data in Excel to export final file.")
        return

    df_export = df_export[
        (df_export["Qty"] != 0) & 
        (df_export["Margin Extended Price"] != 0)
    ]

    if df_export.empty:
        print("Filtered data is empty. Nothing to export.")
        return

    wb_final = Workbook()
    ws_final = wb_final.active
    ws_final.title = "Quote"

    # Copy header including merged cells
    for merged_range in ws_template.merged_cells.ranges:
        ws_final.merge_cells(str(merged_range))

    for row in ws_template.iter_rows(min_row=1, max_row=ws_template.max_row):
        for cell in row:
            new_cell = ws_final.cell(row=cell.row, column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)

    insert_row = ws_template.max_row + 2

    headers = ["Line", "Item Code", "Item Description", "Quantity", "Unit Price", "Extended Price"]
    for col_idx, val in enumerate(headers, start=1):
        ws_final.cell(row=insert_row, column=col_idx, value=val)

    for i, row in enumerate(df_export.itertuples(index=False), start=1):
        ws_final.append([
            i,
            row[0],  # Part Name
            row[1],  # Description
            row[2],  # Qty
            row[8],  # Margin Unit Price
            row[9],  # Margin Extended Price
        ])

    insert_row = ws_final.max_row + 2

    # Copy footer including formatting from Footer.xlsx
    wb_footer = load_workbook("Footer.xlsx")
    ws_footer = wb_footer.active

    for merged_range in ws_footer.merged_cells.ranges:
        ws_final.merge_cells(start_row=insert_row + merged_range.min_row - 1,
                             start_column=merged_range.min_col,
                             end_row=insert_row + merged_range.max_row - 1,
                             end_column=merged_range.max_col)

    for row in ws_footer.iter_rows():
        for cell in row:
            new_cell = ws_final.cell(row=insert_row + cell.row - 1, column=cell.column, value=cell.value)
            copy_cell_style(cell, new_cell)

    final_output_file = "final_quote.xlsx"
    wb_final.save(final_output_file)
    print(f"Final Excel saved as '{final_output_file}'")

# Buttons
ttk.Button(left_frame, text="Save Product", command=save_product).pack(pady=10)
ttk.Button(left_frame, text="Export to Calc Excel", command=export_to_excel).pack(pady=10)
ttk.Button(left_frame, text="Export to Final Excel", command=export_to_final_excel).pack(pady=10)
ttk.Button(left_frame, text="Clear", command=clear_all).pack(pady=10)

root.mainloop()